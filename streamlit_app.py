# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║       PRICE.FUSION — МОНОЛИТНОЕ СТРИМЛИТ ПРИЛОЖЕНИЕ (ВСЁ В 1 ФАЙЛЕ app.py)   ║
#    ⚡ КРАСОЧНЫЙ НЕОНОВЫЙ ДИЗАЙН + ГЛАССМОРФИЗМ + ВЕКТОРНЫЙ ПОИСК МИН. ЦЕНЫ       ║
# ║   Оптимизировано для 300k+ строк на файл | Готово к Деплою                   ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from pathlib import Path
from datetime import datetime
from typing import Optional

# ─────────────────────────────────────────────────────────────────────────────
# 0. КОНФИГУРАЦИЯ СТРАНИЦЫ И ШАПКИ
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Price.Fusion — Агрегатор прайсов",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# 1. КРАСОЧНЫЙ ДИЗАЙН И СТИЛИЗАЦИЯ (НЕОН + ГЛАССМОРФИЗМ + ГРАДИЕНТЫ)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@500;700&display=swap');

    /* ── Глобальный фон с красочным космическим свечением ── */
    .stApp {
        background: 
            radial-gradient(circle at 10% 20%, rgba(99, 102, 241, 0.15) 0%, transparent 40%),
            radial-gradient(circle at 90% 80%, rgba(16, 185, 129, 0.12) 0%, transparent 40%),
            radial-gradient(circle at 50% 50%, rgba(168, 85, 247, 0.1) 0%, transparent 50%),
            linear-gradient(135deg, #090d16 0%, #0d1527 40%, #111a33 100%) !important;
        color: #f1f5f9 !important;
        font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }

    /* ── Сайдбар с неоновой окантовкой ── */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0b1120 0%, #0f172a 100%) !important;
        border-right: 1px solid rgba(99, 102, 241, 0.2) !important;
        box-shadow: 10px 0 30px rgba(0, 0, 0, 0.5);
    }
    [data-testid="stSidebar"] * {
        color: #cbd5e1 !important;
    }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 {
        color: #ffffff !important;
    }
    [data-testid="stToolbar"] { display: none !important; }
    [data-testid="stDecoration"] { display: none !important; }
    header[data-testid="stHeader"] { background: transparent !important; }

    /* ── Красочный Логотип ── */
    .logo-container {
        display: flex;
        align-items: center;
        gap: 0.85rem;
        padding: 0.8rem 1rem;
        background: rgba(255, 255, 255, 0.03);
        border: 1px solid rgba(168, 85, 247, 0.3);
        border-radius: 16px;
        backdrop-filter: blur(12px);
        margin-bottom: 1.5rem;
        box-shadow: 0 8px 32px rgba(168, 85, 247, 0.15);
    }
    .logo-icon {
        width: 44px;
        height: 44px;
        border-radius: 12px;
        background: linear-gradient(135deg, #a855f7 0%, #6366f1 50%, #06b6d4 100%);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.4rem;
        box-shadow: 0 0 20px rgba(168, 85, 247, 0.6);
        animation: pulse-glow 3s infinite alternate;
    }
    @keyframes pulse-glow {
        0% { box-shadow: 0 0 15px rgba(168, 85, 247, 0.5); }
        100% { box-shadow: 0 0 30px rgba(6, 182, 212, 0.8); }
    }
    .logo-title {
        font-weight: 800;
        font-size: 1.15rem;
        background: linear-gradient(135deg, #ffffff 0%, #e2e8f0 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        letter-spacing: -0.02em;
    }
    .logo-sub {
        font-size: 0.65rem;
        color: #a855f7 !important;
        font-family: 'JetBrains Mono', monospace;
        letter-spacing: 0.12em;
        text-transform: uppercase;
        font-weight: 700;
    }

    /* ── Красочные Кнопки Управления ── */
    .stButton > button {
        background: linear-gradient(135deg, #6366f1 0%, #4f46e5 50%, #4338ca 100%) !important;
        color: #ffffff !important;
        border: 1px solid rgba(165, 180, 252, 0.4) !important;
        border-radius: 14px !important;
        padding: 0.65rem 1.4rem !important;
        font-weight: 700 !important;
        font-size: 0.88rem !important;
        letter-spacing: 0.02em !important;
        transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
        box-shadow: 0 4px 20px rgba(99, 102, 241, 0.35) !important;
        width: 100% !important;
    }
    .stButton > button:hover {
        transform: translateY(-2px) scale(1.01) !important;
        box-shadow: 0 8px 30px rgba(99, 102, 241, 0.6) !important;
        background: linear-gradient(135deg, #818cf8 0%, #6366f1 100%) !important;
        border-color: rgba(255, 255, 255, 0.8) !important;
    }
    .stButton > button:active {
        transform: translateY(0) !important;
    }

    /* ── Кнопка Загрузить Демо (Красочный Неоново-Фиолетовый) ── */
    .demo-btn > button {
        background: linear-gradient(135deg, #a855f7 0%, #9333ea 50%, #7e22ce 100%) !important;
        box-shadow: 0 4px 20px rgba(168, 85, 247, 0.4) !important;
        border: 1px solid rgba(233, 213, 255, 0.4) !important;
    }
    .demo-btn > button:hover {
        background: linear-gradient(135deg, #c084fc 0%, #a855f7 100%) !important;
        box-shadow: 0 8px 30px rgba(168, 85, 247, 0.7) !important;
    }

    /* ── Кнопки скачивания (Красочно-Изумрудные) ── */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #10b981 0%, #059669 50%, #047857 100%) !important;
        color: #ffffff !important;
        border: 1px solid rgba(110, 231, 183, 0.4) !important;
        border-radius: 14px !important;
        padding: 0.65rem 1.4rem !important;
        font-weight: 700 !important;
        font-size: 0.88rem !important;
        transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
        box-shadow: 0 4px 20px rgba(16, 185, 129, 0.35) !important;
        width: 100% !important;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px) scale(1.01) !important;
        box-shadow: 0 8px 30px rgba(16, 185, 129, 0.6) !important;
        background: linear-gradient(135deg, #34d399 0%, #10b981 100%) !important;
    }

    /* ── Загрузчик файлов с подсветкой ── */
    [data-testid="stFileUploader"] {
        background: rgba(15, 23, 42, 0.6) !important;
        border: 2px dashed rgba(168, 85, 247, 0.35) !important;
        border-radius: 20px !important;
        padding: 1.4rem !important;
        transition: all 0.3s ease;
        backdrop-filter: blur(10px);
    }
    [data-testid="stFileUploader"]:hover {
        border-color: rgba(6, 182, 212, 0.8) !important;
        background: rgba(15, 23, 42, 0.8) !important;
        box-shadow: 0 0 25px rgba(6, 182, 212, 0.25);
    }
    [data-testid="stFileUploader"] * {
        color: #f1f5f9 !important;
    }
    [data-testid="stFileUploader"] button {
        background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%) !important;
        color: #ffffff !important;
        border: 1px solid rgba(148, 163, 184, 0.3) !important;
        border-radius: 12px !important;
        font-weight: 600 !important;
    }

    /* ── Заголовок и Баннер ── */
    .hero-banner {
        background: linear-gradient(135deg, rgba(30, 41, 59, 0.7) 0%, rgba(15, 23, 42, 0.8) 100%);
        border: 1px solid rgba(168, 85, 247, 0.25);
        border-radius: 24px;
        padding: 2.2rem 2.5rem;
        margin-bottom: 2rem;
        backdrop-filter: blur(16px);
        box-shadow: 0 20px 50px rgba(0, 0, 0, 0.4), inset 0 1px 0 rgba(255, 255, 255, 0.1);
        position: relative;
        overflow: hidden;
    }
    .hero-banner::before {
        content: '';
        position: absolute;
        top: -50%;
        right: -10%;
        width: 300px;
        height: 300px;
        background: radial-gradient(circle, rgba(6, 182, 212, 0.25) 0%, transparent 70%);
        pointer-events: none;
    }
    .hero-title {
        font-size: 2.8rem;
        font-weight: 800;
        background: linear-gradient(135deg, #ffffff 0%, #c084fc 40%, #38bdf8 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        line-height: 1.15;
        letter-spacing: -0.03em;
        margin-bottom: 0.6rem;
    }
    .hero-sub {
        color: #94a3b8 !important;
        font-size: 1.05rem;
        line-height: 1.6;
        max-width: 720px;
    }

    /* ── Карточки шагов (Onboarding) ── */
    .step-card {
        background: rgba(15, 23, 42, 0.6);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 20px;
        padding: 1.6rem;
        height: 100%;
        backdrop-filter: blur(12px);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
    }
    .step-card:hover {
        transform: translateY(-4px);
        border-color: rgba(168, 85, 247, 0.4);
        box-shadow: 0 12px 30px rgba(168, 85, 247, 0.15);
    }
    .step-badge {
        width: 42px;
        height: 42px;
        border-radius: 12px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 800;
        font-size: 1.1rem;
        margin-bottom: 1rem;
    }
    .step-1 { background: linear-gradient(135deg, rgba(99, 102, 241, 0.3), rgba(168, 85, 247, 0.3)); color: #c084fc; border: 1px solid rgba(168, 85, 247, 0.5); }
    .step-2 { background: linear-gradient(135deg, rgba(6, 182, 212, 0.3), rgba(59, 130, 246, 0.3)); color: #38bdf8; border: 1px solid rgba(59, 130, 246, 0.5); }
    .step-3 { background: linear-gradient(135deg, rgba(16, 185, 129, 0.3), rgba(5, 150, 105, 0.3)); color: #34d399; border: 1px solid rgba(16, 185, 129, 0.5); }

    /* ── Красочные KPI Метрики ── */
    .metric-box {
        background: rgba(15, 23, 42, 0.65);
        border-radius: 20px;
        padding: 1.25rem 1.4rem;
        backdrop-filter: blur(12px);
        transition: all 0.25s ease;
        position: relative;
        overflow: hidden;
    }
    .metric-box:hover {
        transform: translateY(-3px);
    }
    .m-blue { border: 1px solid rgba(99, 102, 241, 0.35); box-shadow: 0 10px 25px rgba(99, 102, 241, 0.12); }
    .m-cyan { border: 1px solid rgba(6, 182, 212, 0.35); box-shadow: 0 10px 25px rgba(6, 182, 212, 0.12); }
    .m-purple { border: 1px solid rgba(168, 85, 247, 0.35); box-shadow: 0 10px 25px rgba(168, 85, 247, 0.12); }
    .m-emerald { border: 1px solid rgba(16, 185, 129, 0.35); box-shadow: 0 10px 25px rgba(16, 185, 129, 0.12); }
    .m-amber { border: 1px solid rgba(245, 158, 11, 0.35); box-shadow: 0 10px 25px rgba(245, 158, 11, 0.12); }

    .metric-label {
        font-size: 0.7rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.12em;
        margin-bottom: 0.4rem;
        display: flex;
        align-items: center;
        gap: 0.4rem;
    }
    .ml-blue { color: #818cf8; }
    .ml-cyan { color: #22d3ee; }
    .ml-purple { color: #c084fc; }
    .ml-emerald { color: #34d399; }
    .ml-amber { color: #fbbf24; }

    .metric-val {
        font-size: 1.85rem;
        font-weight: 800;
        color: #ffffff;
        letter-spacing: -0.03em;
        line-height: 1.1;
    }
    .metric-sub {
        font-size: 0.72rem;
        color: #94a3b8;
        margin-top: 0.35rem;
    }

    /* ── Карточки Файлов Источников ── */
    .file-card {
        background: rgba(15, 23, 42, 0.65);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 18px;
        padding: 1.1rem 1.4rem;
        margin-bottom: 0.85rem;
        backdrop-filter: blur(12px);
        transition: all 0.25s ease;
    }
    .file-card:hover {
        border-color: rgba(168, 85, 247, 0.4);
    }
    .fc-ok { border-left: 4px solid #10b981 !important; }
    .fc-warn { border-left: 4px solid #f59e0b !important; }
    .fc-err { border-left: 4px solid #ef4444 !important; }

    /* ── Яркие Бейджи ── */
    .badge {
        display: inline-flex;
        align-items: center;
        padding: 0.25rem 0.75rem;
        border-radius: 9999px;
        font-size: 0.72rem;
        font-weight: 700;
        letter-spacing: 0.04em;
    }
    .b-green { background: rgba(16, 185, 129, 0.18); color: #6ee7b7; border: 1px solid rgba(16, 185, 129, 0.4); }
    .b-red { background: rgba(239, 68, 68, 0.18); color: #fca5a5; border: 1px solid rgba(239, 68, 68, 0.4); }
    .b-yellow { background: rgba(245, 158, 11, 0.18); color: #fcd34d; border: 1px solid rgba(245, 158, 11, 0.4); }
    .b-purple { background: rgba(168, 85, 247, 0.18); color: #e9d5ff; border: 1px solid rgba(168, 85, 247, 0.4); }
    .b-cyan { background: rgba(6, 182, 212, 0.18); color: #a5f3fc; border: 1px solid rgba(6, 182, 212, 0.4); }

    /* ── Разделы и Заголовки ── */
    .sec-head {
        font-size: 0.75rem;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 0.16em;
        color: #a855f7;
        margin-bottom: 0.9rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }

    /* ── Инпуты и Селекты ── */
    input, select, .stSelectbox > div > div, .stTextInput > div > div > input, [data-testid="stTextInputRootElement"] input {
        background: #0f172a !important;
        border: 1px solid rgba(148, 163, 184, 0.25) !important;
        color: #ffffff !important;
        border-radius: 12px !important;
    }
    input::placeholder, .stTextInput input::placeholder, [data-testid="stTextInputRootElement"] input::placeholder {
        color: #64748b !important;
    }
    div[data-baseweb="select"] > div {
        background-color: #0f172a !important;
        color: #ffffff !important;
        border: 1px solid rgba(148, 163, 184, 0.25) !important;
        border-radius: 12px !important;
    }
    div[data-baseweb="select"] * { color: #ffffff !important; }
    div[data-baseweb="menu"] {
        background-color: #0f172a !important;
        border: 1px solid rgba(168, 85, 247, 0.3) !important;
    }
    div[data-baseweb="option"]:hover {
        background-color: #1e293b !important;
    }

    /* ── Табы ── */
    [data-testid="stTabs"] > div:first-child {
        border-bottom: 1px solid rgba(255, 255, 255, 0.1);
        gap: 0.5rem;
    }
    button[data-baseweb="tab"] {
        background: transparent !important;
        color: #94a3b8 !important;
        font-weight: 700;
        font-size: 0.88rem;
        padding: 0.75rem 1.6rem !important;
        border-radius: 12px 12px 0 0 !important;
        transition: all 0.2s;
    }
    button[data-baseweb="tab"]:hover {
        color: #ffffff !important;
        background: rgba(255, 255, 255, 0.05) !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #ffffff !important;
        border-bottom: 3px solid #a855f7 !important;
        background: linear-gradient(180deg, rgba(168, 85, 247, 0.15) 0%, rgba(168, 85, 247, 0.02) 100%) !important;
    }

    /* ── Таблица DataFrame ── */
    div[data-testid="stDataFrame"] {
        background: rgba(15, 23, 42, 0.6) !important;
        border-radius: 16px !important;
        border: 1px solid rgba(255, 255, 255, 0.08) !important;
        backdrop-filter: blur(12px);
    }

    /* ── Expander ── */
    [data-testid="stExpander"] {
        background: rgba(15, 23, 42, 0.5) !important;
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        border-radius: 14px !important;
    }

    /* ── Progress bar ── */
    [data-testid="stProgress"] > div > div {
        background: linear-gradient(90deg, #a855f7, #06b6d4, #10b981) !important;
    }

    hr { border: none; border-top: 1px solid rgba(255, 255, 255, 0.08); margin: 1.8rem 0; }
    ::-webkit-scrollbar { width: 7px; height: 7px; }
    ::-webkit-scrollbar-track { background: #090d16; }
    ::-webkit-scrollbar-thumb { background: rgba(168, 85, 247, 0.4); border-radius: 999px; }
    ::-webkit-scrollbar-thumb:hover { background: rgba(168, 85, 247, 0.7); }
    </style>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# 2. КЛЮЧЕВЫЕ СЛОВА ДЛЯ АВТО-ОПРЕДЕЛЕНИЯ КОЛОНОК
# ─────────────────────────────────────────────────────────────────────────────
ARTICLE_KW = [
    "артикул", "арт", "sku", "код", "кодтовара", "код товара",
    "номер", "part", "oem", "деталь", "catalog", "article",
    "шифр", "партномер", "code", "product code", "item",
    "ean", "upc", "gtin", "mpn", "model", "модель", "номенклатура"
]

BRAND_KW = [
    "бренд", "brand", "производитель", "произв", "марка",
    "фирма", "изготовитель", "maker", "manufacturer",
    "вендор", "make", "producer", "торговаямарка"
]

PRICE_KW = [
    "цена", "price", "стоимость", "прайс", "закуп",
    "розница", "опт", "закупка", "cost", "ррц",
    "rub", "грн", "uah", "usd", "eur", "сумма", "ценасоскидкой",
    "цена со скидкой", "опт цена", "цена закупки",
    "стоимость закупки", "retail", "wholesale",
    "amount", "total", "sum", "net", "gross", "ценаруб"
]

# ─────────────────────────────────────────────────────────────────────────────
# 3. ФУНКЦИИ ОБРАБОТКИ
# ─────────────────────────────────────────────────────────────────────────────

def normalize_header(h: str) -> str:
    """Нормализация заголовка для робастного сопоставления."""
    return re.sub(r"[\s_\-\.\,\/\(\)]+", "", str(h).strip().lower())

def detect_column(columns: list[str], keywords: list[str]) -> Optional[str]:
    """Автоматическое обнаружение колонки по списку ключевых слов."""
    best_col = None
    best_score = -1
    for col in columns:
        norm = normalize_header(col)
        if not norm:
            continue
        for kw in keywords:
            kw_n = normalize_header(kw)
            if norm == kw_n:
                return col
            if norm.startswith(kw_n) and len(kw_n) > best_score:
                best_score = len(kw_n)
                best_col = col
            elif kw_n in norm and len(kw_n) > best_score:
                best_score = len(kw_n)
                best_col = col
    return best_col

def clean_price_vectorized(series: pd.Series) -> pd.Series:
    """
    Высокоскоростная векторизованная очистка цен из любого текстового формата.
    Преобразует строку в float64, игнорируя валютные символы, пробелы и буквы.
    """
    s = series.fillna("").astype(object).astype(str).str.strip()

    # Удаление невидимых и Unicode пробелов
    s = s.str.replace("\u00a0", " ", regex=False)
    s = s.str.replace("\u202f", " ", regex=False)
    s = s.str.replace("\u2009", " ", regex=False)
    s = s.str.replace("\u200b", "", regex=False)

    # Удаление всех обычных пробелов
    s = s.str.replace(r"\s+", "", regex=True)

    # Очистка от знаков валют и суффиксов
    s = s.str.replace(r"[₽$€£¥₴]", "", regex=True)
    s = s.str.replace(r"руб\.?", "", regex=True, case=False)
    s = s.str.replace(r"RUB|rub|USD|usd|EUR|eur|UAH|uah", "", regex=True, case=False)

    # Пустые / нечисловые значения в NaN
    s = s.replace({"": np.nan, "nan": np.nan, "none": np.nan, "None": np.nan, "null": np.nan})

    # Авто-детект десятичных разделителей
    has_comma = s.str.contains(",", na=False)
    has_dot = s.str.contains(r"\.", na=False, regex=True)

    both = has_comma & has_dot
    comma_right = s.str.rfind(",") > s.str.rfind(r"\.")
    mask1 = both & comma_right
    mask2 = both & ~comma_right

    only_comma = has_comma & ~has_dot
    decimal_comma = only_comma & s.str.contains(r",\d{1,2}$", na=False, regex=True)

    result = s.copy()
    result[mask1] = result[mask1].str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
    result[mask2] = result[mask2].str.replace(",", "", regex=False)
    result[decimal_comma] = result[decimal_comma].str.replace(",", ".", regex=False)
    result[only_comma & ~decimal_comma] = result[only_comma & ~decimal_comma].str.replace(",", "", regex=False)

    numeric = pd.to_numeric(result, errors="coerce")
    return numeric.where(numeric > 0, np.nan)

def find_header_row(df_raw: pd.DataFrame) -> int:
    """Умный поиск строки с заголовками таблицы в первых 30 строках."""
    best_idx = 0
    best_score = -1
    for i in range(min(30, len(df_raw))):
        row = df_raw.iloc[i]
        score = 0
        has_art = False
        has_price = False
        for cell in row:
            norm = normalize_header(str(cell))
            for kw in ARTICLE_KW:
                if normalize_header(kw) in norm or norm in normalize_header(kw):
                    has_art = True
                    score += 2
            for kw in PRICE_KW:
                if normalize_header(kw) in norm or norm in normalize_header(kw):
                    has_price = True
                    score += 2
            for kw in BRAND_KW:
                if normalize_header(kw) in norm or norm in normalize_header(kw):
                    score += 1
        if has_art and has_price:
            score += 5
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx

def read_file(uploaded_file) -> Optional[pd.DataFrame]:
    """Безопасное чтение файлов формата xlsx, csv, ods."""
    name = uploaded_file.name.lower()
    try:
        if name.endswith(".csv"):
            for enc in ["utf-8-sig", "utf-8", "cp1251", "latin-1"]:
                try:
                    uploaded_file.seek(0)
                    return pd.read_csv(uploaded_file, encoding=enc, sep=None, engine="python", header=None, dtype=str)
                except Exception:
                    continue
            return None
        elif name.endswith(".ods"):
            uploaded_file.seek(0)
            return pd.read_excel(uploaded_file, engine="odf", header=None, dtype=str)
        else:
            uploaded_file.seek(0)
            return pd.read_excel(uploaded_file, header=None, dtype=str)
    except Exception:
        return None

def parse_price_file(uploaded_file, forced_cols: Optional[dict] = None) -> dict:
    """Парсинг файла, извлечение колонок и сборка очищенного датафрейма."""
    result = {
        "name": uploaded_file.name,
        "status": "error",
        "message": "",
        "col_art": None,
        "col_brand": None,
        "col_price": None,
        "header_row": 0,
        "df_clean": pd.DataFrame(),
        "row_count": 0,
    }

    df_raw = read_file(uploaded_file)
    if df_raw is None or df_raw.empty:
        result["message"] = "Файл пустой или поврежден"
        return result

    hrow = find_header_row(df_raw)
    result["header_row"] = hrow

    headers = df_raw.iloc[hrow].tolist()
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers)]

    df = df_raw.iloc[hrow + 1:].copy()
    df.columns = headers
    df = df.dropna(how="all")

    if df.empty:
        result["message"] = "Нет данных после шапки"
        return result

    # Определение колонок
    col_art = forced_cols.get("art") if forced_cols and forced_cols.get("art") != "—" else detect_column(headers, ARTICLE_KW)
    col_brand = forced_cols.get("brand") if forced_cols and forced_cols.get("brand") != "—" else detect_column(headers, BRAND_KW)
    col_price = forced_cols.get("price") if forced_cols and forced_cols.get("price") != "—" else detect_column(headers, PRICE_KW)

    result["col_art"] = col_art
    result["col_brand"] = col_brand
    result["col_price"] = col_price

    if not col_art or not col_price:
        result["message"] = f"❌ Не найдены обязательные столбцы (Артикул: {bool(col_art)}, Цена: {bool(col_price)})"
        return result

    if col_art not in df.columns or col_price not in df.columns:
        result["message"] = f"❌ Столбцы не найдены в данных ({col_art}, {col_price})"
        return result

    source = Path(uploaded_file.name).stem

    art_series = df[col_art].fillna("").astype(str).str.strip()
    invalid_art = art_series.isin(["", "nan", "none", "NaN", "None", "null", "undefined"]) | art_series.isna()

    price_series = clean_price_vectorized(df[col_price])

    if col_brand and col_brand in df.columns:
        brand_series = df[col_brand].fillna("").astype(str).str.strip()
        brand_series = brand_series.where(~brand_series.isin(["", "nan", "none", "NaN", "None"]), "—")
        brand_series = brand_series.replace({"": "—", np.nan: "—"})
    else:
        brand_series = pd.Series("—", index=df.index)

    valid_mask = ~invalid_art & price_series.notna()

    if valid_mask.sum() == 0:
        result["message"] = "⚠️ Нет строк с валидными артикулами и ценами"
        result["status"] = "warning"
        return result

    df_clean = pd.DataFrame({
        "Артикул": art_series[valid_mask].values,
        "Бренд": brand_series[valid_mask].values,
        "Цена": price_series[valid_mask].values,
        "Источник": source,
    }).reset_index(drop=True)

    result["df_clean"] = df_clean
    result["row_count"] = len(df_clean)
    result["status"] = "ok"
    result["message"] = f"✅ Обработано: {len(df_clean):,} позиций"
    return result

def aggregate_best_prices(df_all: pd.DataFrame) -> pd.DataFrame:
    """
    ПОИСК МИНИМАЛЬНОЙ ЦЕНИ СО ВСЕХ ПРАЙСОВ.
    Сравнивает цены всех поставщиков по каноническому артикулу, выбирает минимум и сбережения.
    """
    if df_all.empty:
        return pd.DataFrame()

    df = df_all.copy()

    df = df[
        df["Артикул"].notna()
        & (df["Артикул"].str.strip() != "")
        & (~df["Артикул"].str.strip().str.lower().isin(["nan", "none", "null", "undefined"]))
    ].copy()

    if df.empty:
        return pd.DataFrame()

    # Дедупликация совпадений в рамках одного источника (берем мин цену внутри прайса)
    df = df.sort_values("Цена").drop_duplicates(subset=["Артикул", "Источник"], keep="first")

    # Нормализованный ключ артикула
    df["_key"] = df["Артикул"].str.upper().str.replace(r"[\s\-_\.\,]+", "", regex=True)

    # Расчет статистики по каждому артикулу
    grp = df.groupby("_key").agg(
        Цена_мин=("Цена", "min"),
        Цена_макс=("Цена", "max"),
        Предложений=("Источник", "nunique"),
    ).reset_index()

    # Точное определение строки с МИНИМАЛЬНОЙ ценой
    idx_min = df.groupby("_key")["Цена"].idxmin()
    df_best = df.loc[idx_min].copy()
    df_best = df_best.merge(grp, on="_key", how="left")
    df_best = df_best.drop(columns=["_key", "Цена"], errors="ignore")
    df_best = df_best.rename(columns={"Цена_мин": "Цена"})

    # Расчет выгоды
    df_best["Экономия_руб"] = (df_best["Цена_макс"] - df_best["Цена"]).round(2)
    df_best["Экономия_%"] = np.where(
        df_best["Цена_макс"] > 0,
        ((df_best["Экономия_руб"] / df_best["Цена_макс"]) * 100).round(1),
        0.0,
    )

    cols = ["Артикул", "Бренд", "Цена", "Источник", "Предложений", "Цена_макс", "Экономия_руб", "Экономия_%"]
    df_best = df_best[[c for c in cols if c in df_best.columns]]
    df_best = df_best.sort_values("Артикул").reset_index(drop=True)
    return df_best

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Генерация стилизованного Excel файла."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Минимальные цены")
        ws = writer.sheets["Минимальные цены"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1e1b4b")
        header_font = Font(bold=True, color="c084fc", size=11)
        thin_border = Border(bottom=Side(style="thin", color="312e81"))
        price_font = Font(bold=True, color="34d399", size=11)

        price_col_idx = None
        for i, col in enumerate(df.columns, 1):
            if col == "Цена":
                price_col_idx = i

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name.upper()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            max_len = max(
                len(str(col_name)),
                *[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, min(ws.max_row + 1, 200))]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
                if price_col_idx and cell.column == price_col_idx:
                    cell.font = price_font
                    if cell.value is not None:
                        cell.number_format = "#,##0.00 \u20bd"

        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A2"
    return output.getvalue()

def to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Генерация CSV файла в UTF-8-BOM."""
    return df.to_csv(index=False, encoding="utf-8-sig", sep=";").encode("utf-8-sig")

# ─────────────────────────────────────────────────────────────────────────────
# 4. СЕССИИ И ИНИЦИАЛИЗАЦИЯ
# ─────────────────────────────────────────────────────────────────────────────
if "parsed_results" not in st.session_state:
    st.session_state.parsed_results = []
if "df_final" not in st.session_state:
    st.session_state.df_final = pd.DataFrame()
if "uploaded_objects" not in st.session_state:
    st.session_state.uploaded_objects = []

# ─────────────────────────────────────────────────────────────────────────────
# 5. КРАСОЧНЫЙ САЙДБАР
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        """
        <div class="logo-container">
          <div class="logo-icon">⚡</div>
          <div>
            <div class="logo-title">PRICE.FUSION</div>
            <div class="logo-sub">Агрегатор Прайсов</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="sec-head">📂 Выбор Прайс-Листов</div>', unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Перетащите файлы сюда",
        type=["xlsx", "xls", "csv", "ods"],
        accept_multiple_files=True,
        help="Загрузите неограниченное число файлов прайс-листов"
    )

    if uploaded_files:
        st.session_state.uploaded_objects = uploaded_files

    st.markdown("<br>", unsafe_allow_html=True)

    c_run, c_clr = st.columns(2)
    with c_run:
        btn_analyze = st.button("▶ Анализ", disabled=not st.session_state.uploaded_objects)
    with c_clr:
        if st.button("✕ Сброс"):
            st.session_state.parsed_results = []
            st.session_state.df_final = pd.DataFrame()
            st.session_state.uploaded_objects = []
            st.rerun()

    st.markdown("---")
    st.markdown('<div class="sec-head">🚀 Быстрый Старт</div>', unsafe_allow_html=True)
    st.caption("Нет своих файлов? Проверьте работу на демо-пакете из 3 поставщиков.")

    st.markdown('<div class="demo-btn">', unsafe_allow_html=True)
    if st.button("⚡ Загрузить Демо Прайсы", use_container_width=True):
        df_demo1 = pd.DataFrame({
            "Артикул товара": ["IP15-128-BLK", "IP15-256-WHT", "SAM-S24-256", "XIA-14-512", "SONY-XM5", "MAC-AIR-M3"],
            "Производитель": ["Apple", "Apple", "Samsung", "Xiaomi", "Sony", "Apple"],
            "Цена закупки": [78900, 89900, 69900, 54900, 34500, 142000]
        })
        df_demo2 = pd.DataFrame({
            "Код SKU": ["IP15-128-BLK", "IP15-256-WHT", "SAM-S24-256", "XIA-14-512", "SONY-XM5", "DYSON-HS05"],
            "Бренд": ["Apple LLC", "Apple", "Samsung Group", "Xiaomi Corp", "Sony", "Dyson"],
            "Опт цена": [76500, 91200, 67800, 56900, 32900, 47900]
        })
        df_demo3 = pd.DataFrame({
            "Артикул": ["IP15-128-BLK", "IP15-256-WHT", "SAM-S24-256", "XIA-14-512", "MAC-AIR-M3", "PS5-SLIM"],
            "Бренд": ["Apple", "Apple", "Samsung", "Xiaomi", "Apple Store", "Sony"],
            "Цена со скидкой": [79000, 88500, 71000, 52900, 139000, 48900]
        })

        def df_to_uploaded(df, filename):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Prices")
            output.seek(0)
            class MockFile(io.BytesIO):
                def __init__(self, val, name):
                    super().__init__(val)
                    self.name = name
            return MockFile(output.getvalue(), filename)

        demo_files = [
            df_to_uploaded(df_demo1, "1_Марвел_Дистрибьюция.xlsx"),
            df_to_uploaded(df_demo2, "2_ОптТорг_Смартфоны.xlsx"),
            df_to_uploaded(df_demo3, "3_Премиум_Импорт.xlsx"),
        ]

        st.session_state.uploaded_objects = demo_files
        results = []
        all_frames = []
        for uf in demo_files:
            res = parse_price_file(uf)
            results.append(res)
            if res["status"] == "ok" and not res["df_clean"].empty:
                all_frames.append(res["df_clean"])

        st.session_state.parsed_results = results
        if all_frames:
            df_all = pd.concat(all_frames, ignore_index=True)
            st.session_state.df_final = aggregate_best_prices(df_all)
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown(
        """
        <div style="font-size:0.75rem;color:#94a3b8;line-height:1.6">
          <b style="color:#c084fc">Автоматический поиск:</b><br>
          • 🔑 Артикул / SKU / OEM<br>
          • 🏷 Бренд / Производитель<br>
          • 💰 Цена / Закупка / Опт
        </div>
        """,
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────────────────────────────────────
# 6. ОБРАБОТКА И АНАЛИЗ ПО КНОПКЕ
# ─────────────────────────────────────────────────────────────────────────────
if 'btn_analyze' not in locals():
    btn_analyze = False

if btn_analyze and st.session_state.uploaded_objects:
    results = []
    all_frames = []

    total_files = len(st.session_state.uploaded_objects)
    prog = st.progress(0, text=f"⚡ Обработка файлов (0/{total_files})...")

    for idx, uf in enumerate(st.session_state.uploaded_objects):
        prog.progress((idx) / total_files, text=f"Анализ {idx+1}/{total_files}: {uf.name}")

        forced = {}
        if st.session_state.get(f"ovr_art_{uf.name}"):
            forced["art"] = st.session_state[f"ovr_art_{uf.name}"]
        if st.session_state.get(f"ovr_brand_{uf.name}"):
            forced["brand"] = st.session_state[f"ovr_brand_{uf.name}"]
        if st.session_state.get(f"ovr_price_{uf.name}"):
            forced["price"] = st.session_state[f"ovr_price_{uf.name}"]

        res = parse_price_file(uf, forced_cols=forced if forced else None)
        results.append(res)
        if res["status"] == "ok" and not res["df_clean"].empty:
            all_frames.append(res["df_clean"])

    prog.progress(1.0, text="✨ Поиск лучших цен...")
    st.session_state.parsed_results = results

    if all_frames:
        df_all = pd.concat(all_frames, ignore_index=True)
        st.session_state.df_final = aggregate_best_prices(df_all)
    else:
        st.session_state.df_final = pd.DataFrame()

    prog.empty()

# ─────────────────────────────────────────────────────────────────────────────
# 7. ГЛАВНЫЙ ИНТЕРФЕЙС И ОНБОРДИНГ
# ─────────────────────────────────────────────────────────────────────────────
df_final: pd.DataFrame = st.session_state.df_final

st.markdown(
    """
    <div class="hero-banner">
      <div class="hero-title">Price.Fusion · Умный Агрегатор Цен</div>
      <div class="hero-sub">
        Загрузите любое количество прайс-листов. Алгоритм автоматически распознает
        <b>Артикул</b>, <b>Бренд</b> и <b>Цену</b>, найдет <b>наименьшую цену</b> по каждому SKU и укажет лучшего поставщика.
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

if df_final.empty and not st.session_state.parsed_results:
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(
            """
            <div class="step-card">
              <div class="step-badge step-1">01</div>
              <div style="font-weight:800;color:#fff;margin-bottom:.4rem;font-size:1.05rem">Умное Распознавание</div>
              <div style="font-size:0.85rem;color:#94a3b8;line-height:1.6">
                Ищет таблицы со сложной структурой и шапками в первых 30 строках файла.
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            """
            <div class="step-card">
              <div class="step-badge step-2">02</div>
              <div style="font-weight:800;color:#fff;margin-bottom:.4rem;font-size:1.05rem">Очистка Цен</div>
              <div style="font-size:0.85rem;color:#94a3b8;line-height:1.6">
                Нормализует любые форматы цен (валюты, пробелы, запятые): <code>1 200,50 ₽</code> → <code>1200.50</code>.
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            """
            <div class="step-card">
              <div class="step-badge step-3">03</div>
              <div style="font-weight:800;color:#fff;margin-bottom:.4rem;font-size:1.05rem">Выбор Лучшей Цены</div>
              <div style="font-size:0.85rem;color:#94a3b8;line-height:1.6">
                Оставляет наименьшую цену по каждому товару с указанием прайса-победителя.
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.info("👈 **Начните работу:** добавьте файлы слева или нажмите **«⚡ Загрузить Демо Прайсы»** для мгновенного теста.", icon="💡")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# 8. КАРТОЧКИ ИСТОЧНИКОВ И НАСТРОЙКА СТОЛБЦОВ
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="sec-head">📁 Загруженные Прайс-Листы</div>', unsafe_allow_html=True)

for res in st.session_state.parsed_results:
    icon = "✅" if res["status"] == "ok" else ("⚠️" if res["status"] == "warning" else "❌")
    cls = "fc-ok" if res["status"] == "ok" else ("fc-warn" if res["status"] == "warning" else "fc-err")

    badge = (
        f'<span class="badge b-green">Успешно · {res["row_count"]:,} позиций</span>'
        if res["status"] == "ok"
        else (
            '<span class="badge b-yellow">ВНИМАНИЕ</span>'
            if res["status"] == "warning"
            else '<span class="badge b-red">ОШИБКА</span>'
        )
    )

    art_b = f'<span class="badge b-purple">Арт: {res["col_art"] or "—"}</span>'
    brand_b = f'<span class="badge b-purple">Бренд: {res["col_brand"] or "—"}</span>'
    price_b = f'<span class="badge b-cyan">Цена: {res["col_price"] or "—"}</span>'

    st.markdown(
        f"""
        <div class="file-card {cls}">
          <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:.6rem">
            <div>
              <span style="font-weight:800;color:#f8fafc;font-size:1rem">{icon} {res['name']}</span>
              &nbsp;{badge}
            </div>
            <div style="display:flex;gap:.4rem;flex-wrap:wrap">
              {art_b} {brand_b} {price_b}
            </div>
          </div>
          <div style="font-size:0.8rem;color:#94a3b8;margin-top:.4rem">{res['message']}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if res["status"] in ("error", "warning") or not res["col_art"] or not res["col_price"]:
        with st.expander(f"🔧 Настроить столбцы вручную для «{res['name']}»"):
            try:
                raw_test = None
                for uf in st.session_state.uploaded_objects:
                    if uf.name == res["name"]:
                        raw_test = read_file(uf)
                        break
                if raw_test is not None:
                    h_idx = res["header_row"]
                    headers_list = ["—"] + [str(x) for x in raw_test.iloc[h_idx].tolist() if pd.notna(x)]
                else:
                    headers_list = ["—"]
            except Exception:
                headers_list = ["—"]

            c_a, c_b, c_p = st.columns(3)
            with c_a:
                st.selectbox("Столбец: Артикул", headers_list, key=f"ovr_art_{res['name']}")
            with c_b:
                st.selectbox("Столбец: Бренд", headers_list, key=f"ovr_brand_{res['name']}")
            with c_p:
                st.selectbox("Столбец: Цена", headers_list, key=f"ovr_price_{res['name']}")
            st.caption("Нажмите кнопку **▶ Анализ** повторно после смены параметров.")

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# 9. ЯРКИЙ KPI ДАШБОРД
# ─────────────────────────────────────────────────────────────────────────────
if not df_final.empty:
    st.markdown('<div class="sec-head">📊 Ключевые Метрики Закупки</div>', unsafe_allow_html=True)

    total_unique = len(df_final)
    total_offers = int(df_final["Предложений"].sum()) if "Предложений" in df_final.columns else 0
    total_savings = df_final["Экономия_руб"].sum() if "Экономия_руб" in df_final.columns else 0
    avg_price = df_final["Цена"].mean()
    total_min_sum = df_final["Цена"].sum()

    m1, m2, m3, m4, m5 = st.columns(5)

    with m1:
        st.markdown(
            f"""
            <div class="metric-box m-purple">
              <div class="metric-label ml-purple">🔑 Уникальных SKU</div>
              <div class="metric-val">{total_unique:,}</div>
              <div class="metric-sub">позиций во всех файлах</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with m2:
        st.markdown(
            f"""
            <div class="metric-box m-cyan">
              <div class="metric-label ml-cyan">📊 Предложений</div>
              <div class="metric-val">{total_offers:,}</div>
              <div class="metric-sub">всего цен от поставщиков</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with m3:
        st.markdown(
            f"""
            <div class="metric-box m-blue">
              <div class="metric-label ml-blue">💰 Ср. Мин. Цена</div>
              <div class="metric-val">{avg_price:,.0f} ₽</div>
              <div class="metric-sub">средняя цена позиции</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with m4:
        st.markdown(
            f"""
            <div class="metric-box m-emerald">
              <div class="metric-label ml-emerald">💵 Бюджет Закупки</div>
              <div class="metric-val">{total_min_sum:,.0f} ₽</div>
              <div class="metric-sub">сумма всех мин. цен</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with m5:
        st.markdown(
            f"""
            <div class="metric-box m-amber">
              <div class="metric-label ml-amber">📉 Сумма Экономии</div>
              <div class="metric-val">{total_savings:,.0f} ₽</div>
              <div class="metric-sub">выгода против макс. цен</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# 10. ТАБЛИЦА, АНАЛИТИКА И ЭКСПОРТ
# ─────────────────────────────────────────────────────────────────────────────
if not df_final.empty:
    tab_table, tab_analysis, tab_sources = st.tabs([
        "📋 Итоговый Прайс", "📈 Аналитика и Графики", "🏢 По Поставщикам"
    ])

    # ── ТАБ 1: ИТОГОВЫЙ ПРАЙС ──
    with tab_table:
        st.markdown('<div class="sec-head">🏆 Сводный Прайс С Минимальной Ценой</div>', unsafe_allow_html=True)

        f_col1, f_col2, f_col3 = st.columns([3, 2, 2])
        with f_col1:
            search_q = st.text_input("Поиск", placeholder="🔍 Поиск по артикулу, бренду или поставщику...", label_visibility="collapsed")
        with f_col2:
            brand_list = ["Все бренды"] + sorted(df_final["Бренд"].dropna().unique().tolist())
            sel_brand_val = st.selectbox("Фильтр по бренду", brand_list, label_visibility="collapsed")
        with f_col3:
            source_list = ["Все источники"] + sorted(df_final["Источник"].dropna().unique().tolist())
            sel_source_val = st.selectbox("Фильтр по источнику", source_list, label_visibility="collapsed")

        df_view = df_final.copy()
        if search_q:
            q = search_q.strip().lower()
            df_view = df_view[
                df_view["Артикул"].str.lower().str.contains(q, na=False) |
                df_view["Бренд"].str.lower().str.contains(q, na=False) |
                df_view["Источник"].str.lower().str.contains(q, na=False)
            ]
        if sel_brand_val != "Все бренды":
            df_view = df_view[df_view["Бренд"] == sel_brand_val]
        if sel_source_val != "Все источники":
            df_view = df_view[df_view["Источник"] == sel_source_val]

        disp_cols = ["Артикул", "Бренд", "Цена", "Источник", "Предложений", "Экономия_руб", "Экономия_%"]
        disp_cols = [c for c in disp_cols if c in df_view.columns]

        st.dataframe(
            df_view[disp_cols],
            use_container_width=True,
            height=500,
            hide_index=True,
            column_config={
                "Артикул": st.column_config.TextColumn("🔑 Артикул", width="medium"),
                "Бренд": st.column_config.TextColumn("🏷 Бренд", width="medium"),
                "Цена": st.column_config.NumberColumn("💰 Мин. цена (₽)", format="%.2f ₽", width="small"),
                "Источник": st.column_config.TextColumn("📁 Лучший Поставщик", width="large"),
                "Предложений": st.column_config.NumberColumn("📊 Источников", width="small"),
                "Экономия_руб": st.column_config.NumberColumn("💚 Экономия (₽)", format="%.2f ₽", width="small"),
                "Экономия_%": st.column_config.NumberColumn("📉 Выгода (%)", format="%.1f%%", width="small"),
            }
        )

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="sec-head">💾 Экспорт Результатов</div>', unsafe_allow_html=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        excel_bytes = to_excel_bytes(df_view[disp_cols])
        csv_bytes = to_csv_bytes(df_view[disp_cols])

        dl_c1, dl_c2 = st.columns(2)
        with dl_c1:
            st.download_button(
                label="📗 Скачать в Excel (.xlsx)",
                data=excel_bytes,
                file_name=f"price_fusion_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl_c2:
            st.download_button(
                label="📄 Скачать в CSV (.csv)",
                data=csv_bytes,
                file_name=f"price_fusion_{timestamp}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ── ТАБ 2: АНАЛИТИКА ──
    with tab_analysis:
        st.markdown('<div class="sec-head">📈 Графики Распределения И Цен</div>', unsafe_allow_html=True)

        gc1, gc2 = st.columns(2)
        with gc1:
            st.markdown("**Распределение минимальных цен**")
            prices_data = df_final["Цена"].dropna()
            p98 = prices_data.quantile(0.98) if len(prices_data) > 5 else prices_data.max()
            clipped = prices_data[prices_data <= p98]
            bins = min(30, max(5, len(clipped) // 10))
            hist = pd.cut(clipped, bins=bins).value_counts().sort_index()
            hist_df = pd.DataFrame({
                "Диапазон": [f"{i.mid:,.0f} ₽" for i in hist.index],
                "Позиций": hist.values,
            })
            st.bar_chart(hist_df.set_index("Диапазон"), color="#a855f7", height=320)

        with gc2:
            st.markdown("**Топ-10 брендов по количеству артикулов**")
            top_brands = df_final[df_final["Бренд"] != "—"]["Бренд"].value_counts().head(10).reset_index()
            top_brands.columns = ["Бренд", "SKU"]
            st.bar_chart(top_brands.set_index("Бренд"), color="#06b6d4", height=320)

    # ── ТАБ 3: ПО ПОСТАВЩИКАМ ──
    with tab_sources:
        st.markdown('<div class="sec-head">🏢 Эффективность Прайс-Листов Поставщиков</div>', unsafe_allow_html=True)

        src_stats = df_final.groupby("Источник").agg(
            Побед=("Источник", "count"),
            Мин_цена=("Цена", "min"),
            Средн_цена=("Цена", "mean"),
            Макс_цена=("Цена", "max"),
            Сумма_мин=("Цена", "sum"),
            Экономия=("Экономия_руб", "sum"),
        ).round(2).sort_values("Побед", ascending=False).reset_index()

        st.dataframe(
            src_stats,
            use_container_width=True,
            height=380,
            hide_index=True,
            column_config={
                "Источник": st.column_config.TextColumn("📁 Название Прайса / Поставщик", width="large"),
                "Побед": st.column_config.NumberColumn("🏆 Побед (мин. цена)", help="Сколько раз этот прайс дал наименьшую цену"),
                "Мин_цена": st.column_config.NumberColumn("Min ₽", format="%.0f ₽"),
                "Средн_цена": st.column_config.NumberColumn("Ср. ₽", format="%.0f ₽"),
                "Макс_цена": st.column_config.NumberColumn("Max ₽", format="%.0f ₽"),
                "Сумма_мин": st.column_config.NumberColumn("💵 Общий бюджет ₽", format="%.0f ₽"),
                "Экономия": st.column_config.NumberColumn("💚 Принесенная Экономия ₽", format="%.0f ₽"),
            }
        )

# ─────────────────────────────────────────────────────────────────────────────
# 11. ПОДВАЛ (FOOTER)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown(
    """
    <hr>
    <div style="text-align:center;font-size:0.78rem;color:#64748b;padding:1rem 0">
      ⚡ <b>Price.Fusion</b> &nbsp;·&nbsp; Монолитный Streamlit App для Деплоя &nbsp;·&nbsp; Все стили и код в одном файле <code>app.py</code>
    </div>
    """,
    unsafe_allow_html=True,
)
